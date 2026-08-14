import csv
import io
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook

from app.db import (
    category_options,
    confirm_import_batch,
    create_import_batch,
    create_project,
    dashboard_summary,
    factory_lookup,
    factory_lookup_by_name,
    get_import_batch,
    get_project,
    init_db,
    list_factories,
    list_projects,
    project_exists_by_code,
    project_exists_by_title_factory,
    project_filters,
    status_options,
    today_iso,
    update_project,
)

BASE_DIR = Path(__file__).resolve().parent.parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))

app = FastAPI(title="Cost Reduction Platform Demo")
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")

ALLOWED_EXTENSIONS = {".csv", ".xlsx"}
REQUIRED_HEADERS = [
    "project_code",
    "title",
    "factory_name",
    "category",
    "status",
    "owner",
    "estimated_savings",
    "currency",
    "start_date",
    "target_date",
    "description",
]


@app.on_event("startup")
def startup() -> None:
    init_db()


def render(request: Request, template_name: str, context: dict[str, Any]) -> HTMLResponse:
    base_context = {
        "request": request,
        "factories": list_factories(),
        "filters": project_filters(),
        "today": today_iso(),
    }
    base_context.update(context)
    return templates.TemplateResponse(template_name, base_context)


def parse_money(value: str) -> float:
    cleaned = str(value).replace(",", "").strip()
    amount = float(cleaned)
    if amount < 0:
        raise ValueError("Estimated savings must be zero or positive.")
    return amount


def parse_date(value: str, field_label: str) -> str:
    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d").date().isoformat()
    except ValueError as exc:
        raise ValueError(f"{field_label} must use YYYY-MM-DD format.") from exc


def collect_form_payload(
    project_code: str,
    title: str,
    factory_id: int,
    category: str,
    status: str,
    owner: str,
    estimated_savings: str,
    currency: str,
    start_date: str,
    target_date: str,
    description: str,
) -> tuple[dict[str, Any], list[str]]:
    errors: list[str] = []
    payload = {
        "project_code": project_code.strip(),
        "title": title.strip(),
        "factory_id": factory_id,
        "category": category.strip(),
        "status": status.strip(),
        "owner": owner.strip(),
        "currency": currency.strip().upper() or "CNY",
        "description": description.strip(),
    }
    for field, label in [
        ("project_code", "Project code"),
        ("title", "Title"),
        ("owner", "Owner"),
        ("description", "Description"),
    ]:
        if not payload[field]:
            errors.append(f"{label} is required.")
    if payload["category"] not in category_options():
        errors.append("Category is invalid.")
    if payload["status"] not in status_options():
        errors.append("Status is invalid.")
    try:
        payload["estimated_savings"] = parse_money(estimated_savings)
    except ValueError as exc:
        errors.append(str(exc))
    try:
        payload["start_date"] = parse_date(start_date, "Start date")
    except ValueError as exc:
        errors.append(str(exc))
    try:
        payload["target_date"] = parse_date(target_date, "Target date")
    except ValueError as exc:
        errors.append(str(exc))
    if not errors and payload["target_date"] < payload["start_date"]:
        errors.append("Target date must be on or after start date.")
    return payload, errors


def normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def parse_upload(filename: str, content: bytes) -> list[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise ValueError("Only CSV and XLSX files are supported.")
    if suffix == ".csv":
        stream = io.StringIO(content.decode("utf-8-sig"))
        reader = csv.DictReader(stream)
        rows = []
        for row in reader:
            rows.append({normalize_header(k): (v or "").strip() for k, v in row.items()})
        return rows

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return []
    headers = [normalize_header(cell) for cell in values[0]]
    rows = []
    for row_values in values[1:]:
        row = {}
        for index, header in enumerate(headers):
            row[header] = "" if index >= len(row_values) or row_values[index] is None else str(row_values[index]).strip()
        rows.append(row)
    return rows


def validate_import_rows(raw_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, int]]:
    factories_by_name = factory_lookup_by_name()
    seen_codes: set[str] = set()
    seen_title_factory: set[tuple[str, int]] = set()
    preview_rows: list[dict[str, Any]] = []
    summary = {"total": 0, "valid": 0, "invalid": 0}

    for index, row in enumerate(raw_rows, start=2):
        errors: list[str] = []
        normalized = {header: str(row.get(header, "")).strip() for header in REQUIRED_HEADERS}
        if any(header not in row for header in REQUIRED_HEADERS):
            missing_headers = [header for header in REQUIRED_HEADERS if header not in row]
            raise ValueError(f"Missing required headers: {', '.join(missing_headers)}")

        factory_name_key = normalized["factory_name"].lower()
        factory = factories_by_name.get(factory_name_key)
        if not factory:
            errors.append("Factory name must be one of Factory Alpha, Factory Beta, or Factory Gamma.")
        if normalized["category"] not in category_options():
            errors.append("Category is invalid.")
        if normalized["status"] not in status_options():
            errors.append("Status is invalid.")
        for field in REQUIRED_HEADERS:
            if not normalized[field]:
                errors.append(f"{field} is required.")

        parsed_amount = None
        parsed_start = normalized["start_date"]
        parsed_target = normalized["target_date"]
        if normalized["estimated_savings"]:
            try:
                parsed_amount = parse_money(normalized["estimated_savings"])
            except ValueError as exc:
                errors.append(str(exc))
        if normalized["start_date"]:
            try:
                parsed_start = parse_date(normalized["start_date"], "Start date")
            except ValueError as exc:
                errors.append(str(exc))
        if normalized["target_date"]:
            try:
                parsed_target = parse_date(normalized["target_date"], "Target date")
            except ValueError as exc:
                errors.append(str(exc))
        if normalized["start_date"] and normalized["target_date"] and parsed_target < parsed_start:
            errors.append("Target date must be on or after start date.")

        factory_id = int(factory["id"]) if factory else None
        title_key = normalized["title"].strip().lower()
        code_key = normalized["project_code"].strip().upper()
        if code_key in seen_codes:
            errors.append("Duplicate project code found within the import file.")
        if factory_id and (title_key, factory_id) in seen_title_factory:
            errors.append("Duplicate title found within the same factory in the import file.")
        if code_key and project_exists_by_code(code_key):
            errors.append("Project code already exists in the database.")
        if factory_id and normalized["title"] and project_exists_by_title_factory(normalized["title"], factory_id):
            errors.append("A project with the same title already exists in this factory.")

        if code_key:
            seen_codes.add(code_key)
        if factory_id and title_key:
            seen_title_factory.add((title_key, factory_id))

        normalized_payload = {
            "project_code": code_key,
            "title": normalized["title"],
            "title_normalized": " ".join(normalized["title"].lower().split()),
            "factory_id": factory_id,
            "category": normalized["category"],
            "status": normalized["status"],
            "owner": normalized["owner"],
            "estimated_savings": parsed_amount,
            "currency": normalized["currency"].upper() or "CNY",
            "start_date": parsed_start,
            "target_date": parsed_target,
            "description": normalized["description"],
        }
        can_import = not errors
        preview_rows.append(
            {
                "row_number": index,
                "data": normalized,
                "errors": errors,
                "can_import": can_import,
                "normalized": normalized_payload,
            }
        )
        summary["total"] += 1
        summary["valid" if can_import else "invalid"] += 1

    return preview_rows, summary


@app.get("/", response_class=HTMLResponse)
def dashboard(request: Request, factory_id: int | None = None) -> HTMLResponse:
    summary = dashboard_summary(factory_id=factory_id)
    return render(request, "dashboard.html", {"summary": summary, "selected_factory_id": factory_id})


@app.get("/projects", response_class=HTMLResponse)
def project_index(
    request: Request,
    q: str = "",
    factory_id: str = "",
    status: str = "",
    category: str = "",
) -> HTMLResponse:
    filters = {"q": q, "factory_id": factory_id, "status": status, "category": category}
    return render(
        request,
        "projects.html",
        {"projects": list_projects(filters), "query": filters},
    )


@app.get("/projects/new", response_class=HTMLResponse)
def project_new(request: Request) -> HTMLResponse:
    return render(
        request,
        "project_form.html",
        {
            "mode": "create",
            "errors": [],
            "form": {
                "currency": "CNY",
                "start_date": today_iso(),
                "target_date": today_iso(),
            },
        },
    )


@app.post("/projects/new", response_class=HTMLResponse)
def project_create(
    request: Request,
    project_code: str = Form(...),
    title: str = Form(...),
    factory_id: int = Form(...),
    category: str = Form(...),
    status: str = Form(...),
    owner: str = Form(...),
    estimated_savings: str = Form(...),
    currency: str = Form(...),
    start_date: str = Form(...),
    target_date: str = Form(...),
    description: str = Form(...),
) -> HTMLResponse:
    payload, errors = collect_form_payload(
        project_code,
        title,
        factory_id,
        category,
        status,
        owner,
        estimated_savings,
        currency,
        start_date,
        target_date,
        description,
    )
    if not errors and project_exists_by_code(payload["project_code"]):
        errors.append("Project code already exists.")
    if not errors and project_exists_by_title_factory(payload["title"], payload["factory_id"]):
        errors.append("A project with the same title already exists in this factory.")
    if errors:
        return render(request, "project_form.html", {"mode": "create", "errors": errors, "form": payload})
    project_id = create_project(payload)
    return RedirectResponse(url=f"/projects/{project_id}", status_code=303)


@app.get("/projects/{project_id}", response_class=HTMLResponse)
def project_detail(request: Request, project_id: int) -> HTMLResponse:
    project = get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return render(request, "project_detail.html", {"project": project})


@app.get("/projects/{project_id}/edit", response_class=HTMLResponse)
def project_edit(request: Request, project_id: int) -> HTMLResponse:
    project = get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return render(request, "project_form.html", {"mode": "edit", "errors": [], "form": project, "project_id": project_id})


@app.post("/projects/{project_id}/edit", response_class=HTMLResponse)
def project_update(
    request: Request,
    project_id: int,
    project_code: str = Form(...),
    title: str = Form(...),
    factory_id: int = Form(...),
    category: str = Form(...),
    status: str = Form(...),
    owner: str = Form(...),
    estimated_savings: str = Form(...),
    currency: str = Form(...),
    start_date: str = Form(...),
    target_date: str = Form(...),
    description: str = Form(...),
) -> HTMLResponse:
    payload, errors = collect_form_payload(
        project_code,
        title,
        factory_id,
        category,
        status,
        owner,
        estimated_savings,
        currency,
        start_date,
        target_date,
        description,
    )
    if not errors and project_exists_by_code(payload["project_code"], exclude_id=project_id):
        errors.append("Project code already exists.")
    if not errors and project_exists_by_title_factory(payload["title"], payload["factory_id"], exclude_id=project_id):
        errors.append("A project with the same title already exists in this factory.")
    if errors:
        return render(
            request,
            "project_form.html",
            {"mode": "edit", "errors": errors, "form": payload, "project_id": project_id},
        )
    update_project(project_id, payload)
    return RedirectResponse(url=f"/projects/{project_id}", status_code=303)


@app.get("/imports", response_class=HTMLResponse)
def import_page(request: Request) -> HTMLResponse:
    return render(request, "imports.html", {})


@app.post("/imports/preview", response_class=HTMLResponse)
async def import_preview(request: Request, file: UploadFile = File(...)) -> HTMLResponse:
    filename = file.filename or ""
    content = await file.read()
    try:
        rows = parse_upload(filename, content)
        preview_rows, summary = validate_import_rows(rows)
        batch_id = create_import_batch(filename, {"rows": preview_rows, "summary": summary})
        return render(
            request,
            "import_preview.html",
            {
                "batch": get_import_batch(batch_id),
                "factory_map": factory_lookup(),
            },
        )
    except ValueError as exc:
        return render(request, "imports.html", {"upload_error": str(exc)})


@app.post("/imports/{batch_id}/confirm", response_class=HTMLResponse)
def import_confirm(request: Request, batch_id: str) -> HTMLResponse:
    batch = get_import_batch(batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="Import batch not found")
    if batch["summary"]["invalid"] > 0:
        return render(
            request,
            "import_preview.html",
            {
                "batch": batch,
                "factory_map": factory_lookup(),
                "confirm_error": "This batch still contains invalid rows. Please fix the file and upload again.",
            },
        )
    inserted = confirm_import_batch(batch_id)
    return render(
        request,
        "import_preview.html",
        {
            "batch": get_import_batch(batch_id),
            "factory_map": factory_lookup(),
            "confirm_success": f"Imported {inserted} synthetic projects successfully.",
        },
    )
